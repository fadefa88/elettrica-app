#!/usr/bin/env python3
"""Audit Motornet car catalog data quality and export a slim Excel workbook.

The Excel export is intentionally limited to fields used by the frontend for car
selection, display, TCO calculations, tax/superbollo logic, and traceability.
Large raw/debug fields such as specs_raw are not exported. Images are ignored
in issue generation.
"""
from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


EXCEL_COLUMNS = [
    "issues",
    "id",
    "category",
    "brand",
    "model",
    "version",
    "fuel",
    "price_eur",
    "power_kw",
    "power_cv",
    "consumption_l_100km",
    "consumption_kg_100km",
    "consumption_kwh_100km",
    "battery_kwh",
    "range_wltp_km",
    "emissions_g_km",
    "image_url",
    "source_url",
]

FUEL_LABELS = {
    "E": "elettrica",
    "EH": "elettrica_idrogeno",
    "B": "benzina",
    "D": "diesel",
    "IB": "ibrida_benzina",
    "ID": "ibrida_diesel",
    "G": "gpl",
    "IG": "ibrida_gpl",
    "M": "metano",
    "IM": "ibrida_metano",
}

ELECTRIC_FUELS = {"elettrica", "elettrica_idrogeno"}
GAS_FUELS = {"metano", "ibrida_metano"}
THERMAL_FUELS = {
    "benzina",
    "diesel",
    "gpl",
    "metano",
    "ibrida_benzina",
    "ibrida_diesel",
    "ibrida_gpl",
    "ibrida_metano",
}


@dataclass(frozen=True)
class AuditLimits:
    ev_consumption_min: float = 7
    ev_consumption_max: float = 40
    ev_battery_min: float = 5
    ev_battery_max: float = 250
    ev_range_min: float = 30
    ev_range_max: float = 1000
    thermal_consumption_l_min: float = 1
    thermal_consumption_l_max: float = 30
    thermal_consumption_kg_min: float = 1
    thermal_consumption_kg_max: float = 15
    emissions_min: float = 1
    emissions_max: float = 500
    price_min: float = 1000
    price_max: float = 600000
    power_kw_min: float = 10
    power_kw_max: float = 900
    power_cv_min: float = 10
    power_cv_max: float = 1300


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def parse_number(value: Any) -> float | None:
    if value is None or value == "" or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        n = float(value)
        return n if math.isfinite(n) and n > 0 else None
    text = clean_text(value)
    if not text:
        return None
    normalized = text.replace(" ", "")
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    else:
        normalized = normalized.replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", normalized)
    if not match:
        return None
    n = float(match.group(0))
    return n if math.isfinite(n) and n > 0 else None


def excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def first_text(car: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = clean_text(car.get(key))
        if value:
            return value
    return ""


def first_number(car: dict[str, Any], *keys: str) -> float | None:
    for key in keys:
        value = parse_number(car.get(key))
        if value is not None:
            return value
    return None


def fuel_of(car: dict[str, Any]) -> str:
    fuel = clean_text(car.get("fuel")).lower()
    if fuel:
        return fuel
    code = clean_text(car.get("fuel_code")).upper()
    if code in FUEL_LABELS:
        return FUEL_LABELS[code]
    return clean_text(car.get("fuel_original")).lower()


def category_of(car: dict[str, Any]) -> str:
    category = clean_text(car.get("category")).lower()
    if category in {"electric", "thermal"}:
        return category
    fuel = fuel_of(car)
    return "electric" if fuel in ELECTRIC_FUELS or "elettr" in fuel else "thermal"


def car_image_url(car: dict[str, Any]) -> str:
    return first_text(car, "image_url", "image_local_path", "image_source_url")


def source_url(car: dict[str, Any]) -> str:
    return first_text(car, "source_url", "motornet_detail_url")


def add_issue(issues: list[str], label: str, value: Any) -> None:
    if isinstance(value, str):
        if not clean_text(value):
            issues.append(f"missing {label}")
    elif value is None or value == "":
        issues.append(f"missing {label}")


def range_issue(issues: list[str], label: str, value: float | None, unit: str, min_value: float, max_value: float) -> None:
    if value is None:
        issues.append(f"missing {label}")
    elif value < min_value:
        issues.append(f"too low {label}: {value:g} {unit} < {min_value:g}")
    elif value > max_value:
        issues.append(f"too high {label}: {value:g} {unit} > {max_value:g}")


def issues_for(car: dict[str, Any], limits: AuditLimits) -> list[str]:
    issues: list[str] = []
    category = category_of(car)
    fuel = fuel_of(car)

    add_issue(issues, "id", first_text(car, "id"))
    add_issue(issues, "brand", first_text(car, "brand"))
    add_issue(issues, "model", first_text(car, "model"))
    add_issue(issues, "version", first_text(car, "version", "powertrain"))
    add_issue(issues, "fuel", fuel)
    add_issue(issues, "source_url", source_url(car))

    range_issue(issues, "price_eur", first_number(car, "price_eur"), "€", limits.price_min, limits.price_max)
    range_issue(issues, "power_kw", first_number(car, "power_kw"), "kW", limits.power_kw_min, limits.power_kw_max)
    range_issue(issues, "power_cv", first_number(car, "power_cv"), "CV", limits.power_cv_min, limits.power_cv_max)

    if category == "electric":
        range_issue(
            issues,
            "consumption_kwh_100km",
            first_number(car, "consumption_kwh_100km"),
            "kWh/100 km",
            limits.ev_consumption_min,
            limits.ev_consumption_max,
        )
        range_issue(issues, "battery_kwh", first_number(car, "battery_kwh"), "kWh", limits.ev_battery_min, limits.ev_battery_max)
        range_issue(issues, "range_wltp_km", first_number(car, "range_wltp_km"), "km", limits.ev_range_min, limits.ev_range_max)
    else:
        if fuel not in THERMAL_FUELS:
            issues.append(f"unknown thermal fuel: {fuel or '-'}")
        if fuel in GAS_FUELS:
            range_issue(
                issues,
                "consumption_kg_100km",
                first_number(car, "consumption_kg_100km"),
                "kg/100 km",
                limits.thermal_consumption_kg_min,
                limits.thermal_consumption_kg_max,
            )
        else:
            range_issue(
                issues,
                "consumption_l_100km",
                first_number(car, "consumption_l_100km"),
                "l/100 km",
                limits.thermal_consumption_l_min,
                limits.thermal_consumption_l_max,
            )
        range_issue(issues, "emissions_g_km", first_number(car, "emissions_g_km"), "g/km", limits.emissions_min, limits.emissions_max)

    return issues


def row_for(car: dict[str, Any], limits: AuditLimits) -> dict[str, Any]:
    issues = issues_for(car, limits)
    return {
        "issues": " | ".join(issues) if issues else "tutto ok",
        "id": first_text(car, "id"),
        "category": category_of(car),
        "brand": first_text(car, "brand"),
        "model": first_text(car, "model"),
        "version": first_text(car, "version", "powertrain"),
        "fuel": fuel_of(car),
        "price_eur": first_number(car, "price_eur") or "",
        "power_kw": first_number(car, "power_kw") or "",
        "power_cv": first_number(car, "power_cv") or "",
        "consumption_l_100km": first_number(car, "consumption_l_100km") or "",
        "consumption_kg_100km": first_number(car, "consumption_kg_100km") or "",
        "consumption_kwh_100km": first_number(car, "consumption_kwh_100km") or "",
        "battery_kwh": first_number(car, "battery_kwh") or "",
        "range_wltp_km": first_number(car, "range_wltp_km") or "",
        "emissions_g_km": first_number(car, "emissions_g_km") or "",
        "image_url": car_image_url(car),
        "source_url": source_url(car),
    }


def load_cars(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    cars = payload.get("cars")
    if not isinstance(cars, list):
        raise SystemExit("Catalog JSON must be an object with cars[]")
    return [car for car in cars if isinstance(car, dict)]


def write_excel(rows: list[dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "catalog"
    ws.append(EXCEL_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")

    for row in rows:
        ws.append([excel_value(row.get(col)) for col in EXCEL_COLUMNS])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "issues": 42,
        "id": 28,
        "brand": 20,
        "model": 48,
        "version": 56,
        "fuel": 18,
        "image_url": 48,
        "source_url": 60,
    }
    for idx, column in enumerate(EXCEL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(column, 16)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_csv(rows: list[dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=EXCEL_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_summary(rows: list[dict[str, Any]], path: Path, limits: AuditLimits) -> None:
    total = len(rows)
    ok = sum(1 for row in rows if row.get("issues") == "tutto ok")
    categories = Counter(row.get("category") or "unknown" for row in rows)
    fuels = Counter(row.get("fuel") or "unknown" for row in rows)
    issue_counter: Counter[str] = Counter()
    for row in rows:
        if row.get("issues") == "tutto ok":
            continue
        for issue in str(row.get("issues") or "").split(" | "):
            if issue:
                issue_counter[issue] += 1

    lines = [
        "# Motornet catalog audit",
        "",
        f"- Total cars: {total}",
        f"- OK rows: {ok}",
        f"- Rows with issues: {total - ok}",
        "",
        "## Audit thresholds",
        "",
        f"- EV consumption: {limits.ev_consumption_min:g}-{limits.ev_consumption_max:g} kWh/100 km",
        f"- EV battery: {limits.ev_battery_min:g}-{limits.ev_battery_max:g} kWh",
        f"- EV WLTP range: {limits.ev_range_min:g}-{limits.ev_range_max:g} km",
        f"- Thermal consumption: {limits.thermal_consumption_l_min:g}-{limits.thermal_consumption_l_max:g} l/100 km",
        f"- Methane consumption: {limits.thermal_consumption_kg_min:g}-{limits.thermal_consumption_kg_max:g} kg/100 km",
        f"- CO2 emissions: {limits.emissions_min:g}-{limits.emissions_max:g} g/km",
        "",
        "## Categories",
        "",
    ]
    lines.extend(f"- {k}: {v}" for k, v in categories.most_common())
    lines.extend(["", "## Fuels", ""])
    lines.extend(f"- {k}: {v}" for k, v in fuels.most_common())
    lines.extend(["", "## Top issues", ""])
    if issue_counter:
        lines.extend(f"- {k}: {v}" for k, v in issue_counter.most_common(50))
    else:
        lines.append("- none")
    lines.extend([
        "",
        "## Output files",
        "",
        "- motornet_catalog_audit.xlsx: curated workbook for manual cleanup",
        "- motornet_catalog_audit.csv: same data in CSV format",
        "- quality_report.md: markdown summary",
        "",
        "Images are intentionally ignored in issue generation.",
    ])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_limits(args: argparse.Namespace) -> AuditLimits:
    return AuditLimits(
        ev_consumption_min=args.ev_consumption_min,
        ev_consumption_max=args.ev_consumption_max,
        ev_battery_min=args.ev_battery_min,
        ev_battery_max=args.ev_battery_max,
        ev_range_min=args.ev_range_min,
        ev_range_max=args.ev_range_max,
        thermal_consumption_l_min=args.thermal_consumption_l_min,
        thermal_consumption_l_max=args.thermal_consumption_l_max,
        thermal_consumption_kg_min=args.thermal_consumption_kg_min,
        thermal_consumption_kg_max=args.thermal_consumption_kg_max,
        emissions_min=args.emissions_min,
        emissions_max=args.emissions_max,
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit Motornet catalog and export a slim Excel workbook.")
    parser.add_argument("--catalog", default="data/cars_motornet.json", help="Motornet JSON catalog")
    parser.add_argument("--out-dir", default="reports/motornet-quality", help="Output directory")
    parser.add_argument("--fail-on-issues", action="store_true", help="Exit with code 1 if anomalies are found")
    parser.add_argument("--ev-consumption-min", type=float, default=AuditLimits.ev_consumption_min)
    parser.add_argument("--ev-consumption-max", type=float, default=AuditLimits.ev_consumption_max)
    parser.add_argument("--ev-battery-min", type=float, default=AuditLimits.ev_battery_min)
    parser.add_argument("--ev-battery-max", type=float, default=AuditLimits.ev_battery_max)
    parser.add_argument("--ev-range-min", type=float, default=AuditLimits.ev_range_min)
    parser.add_argument("--ev-range-max", type=float, default=AuditLimits.ev_range_max)
    parser.add_argument("--thermal-consumption-l-min", type=float, default=AuditLimits.thermal_consumption_l_min)
    parser.add_argument("--thermal-consumption-l-max", type=float, default=AuditLimits.thermal_consumption_l_max)
    parser.add_argument("--thermal-consumption-kg-min", type=float, default=AuditLimits.thermal_consumption_kg_min)
    parser.add_argument("--thermal-consumption-kg-max", type=float, default=AuditLimits.thermal_consumption_kg_max)
    parser.add_argument("--emissions-min", type=float, default=AuditLimits.emissions_min)
    parser.add_argument("--emissions-max", type=float, default=AuditLimits.emissions_max)
    args = parser.parse_args()

    catalog_path = Path(args.catalog)
    out_dir = Path(args.out_dir)
    limits = build_limits(args)
    cars = load_cars(catalog_path)
    rows = [row_for(car, limits) for car in cars]

    out_dir.mkdir(parents=True, exist_ok=True)
    write_excel(rows, out_dir / "motornet_catalog_audit.xlsx")
    write_csv(rows, out_dir / "motornet_catalog_audit.csv")
    write_summary(rows, out_dir / "quality_report.md", limits)
    # Backward-compatible filename used by older runs/docs.
    write_summary(rows, out_dir / "motornet_quality_summary.md", limits)

    issue_rows = sum(1 for row in rows if row.get("issues") != "tutto ok")
    print(f"Cars audited: {len(rows)}")
    print(f"Rows with issues: {issue_rows}")
    print(f"Excel: {out_dir / 'motornet_catalog_audit.xlsx'}")
    print(f"Summary: {out_dir / 'quality_report.md'}")
    if args.fail_on_issues and issue_rows:
        print("Failing because --fail-on-issues is set and anomalies were found.")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
