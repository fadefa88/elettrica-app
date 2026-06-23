#!/usr/bin/env python3
"""Convert a cleaned Motornet Excel workbook back to a slim cars_motornet.json.

The output JSON intentionally contains only fields used by the frontend.
Large scraping/debug fields such as specs_raw, image_source_url, image_bytes,
source_site, fuel_code, fuel_original, year, powertrain and motornet_detail_url
are not written.
"""
from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SLIM_FIELDS = [
    "id",
    "category",
    "brand",
    "model",
    "version",
    "fuel",
    "price_eur",
    "power_kw",
    "power_cv",
    "consumption_l_100km",
    "consumption_kg_100km",
    "consumption_kwh_100km",
    "battery_kwh",
    "range_wltp_km",
    "emissions_g_km",
    "image_url",
    "source_url",
]

TEXT_FIELDS = {"id", "category", "brand", "model", "version", "fuel", "image_url", "source_url"}
NUMERIC_FIELDS = {
    "price_eur",
    "power_kw",
    "power_cv",
    "consumption_l_100km",
    "consumption_kg_100km",
    "consumption_kwh_100km",
    "battery_kwh",
    "range_wltp_km",
    "emissions_g_km",
}
IGNORED_FIELDS = {"issues", "status", "notes"}

FUEL_LABELS = {
    "E": "elettrica",
    "EH": "elettrica_idrogeno",
    "B": "benzina",
    "D": "diesel",
    "IB": "ibrida_benzina",
    "ID": "ibrida_diesel",
    "G": "gpl",
    "IG": "ibrida_gpl",
    "M": "metano",
    "IM": "ibrida_metano",
}


def clean(value: Any) -> str:
    return " ".join(str(value or "").split())


def parse_number(value: Any) -> float | int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        n = float(value)
        if not math.isfinite(n):
            return None
        return int(n) if n.is_integer() else n
    text = clean(value).replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        n = float(text)
    except ValueError:
        return None
    if not math.isfinite(n):
        return None
    return int(n) if n.is_integer() else n


def load_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {sheet_name}. Available: {', '.join(wb.sheetnames)}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(h) for h in rows[0]]
    out: list[dict[str, Any]] = []
    for raw in rows[1:]:
        row = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers)) if headers[i]}
        if clean(row.get("id")):
            out.append(row)
    return out


def load_base_json(path: Path | None) -> dict[str, Any]:
    if path is None:
        return {"source": "motornet.it", "status": "ok", "schema": "cars_motornet_slim_v1", "cars": []}
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict) or not isinstance(payload.get("cars"), list):
        raise SystemExit("Base JSON must be an object with cars[]")
    return payload


def fuel_of(car: dict[str, Any]) -> str:
    fuel = clean(car.get("fuel")).lower()
    if fuel:
        return fuel
    code = clean(car.get("fuel_code")).upper()
    if code in FUEL_LABELS:
        return FUEL_LABELS[code]
    return clean(car.get("fuel_original")).lower()


def category_of(car: dict[str, Any]) -> str:
    category = clean(car.get("category")).lower()
    if category in {"electric", "thermal"}:
        return category
    fuel = fuel_of(car)
    return "electric" if "elettr" in fuel else "thermal"


def first_text(car: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = clean(car.get(key))
        if value:
            return value
    return ""


def first_number(car: dict[str, Any], *keys: str) -> float | int | None:
    for key in keys:
        value = parse_number(car.get(key))
        if value is not None:
            return value
    return None


def apply_excel_row(base: dict[str, Any], row: dict[str, Any], clear_empty: bool) -> dict[str, Any]:
    car = dict(base)
    for field in TEXT_FIELDS:
        if field not in row or field in IGNORED_FIELDS:
            continue
        value = clean(row.get(field))
        if value or clear_empty:
            if value:
                car[field] = value
            else:
                car.pop(field, None)
    for field in NUMERIC_FIELDS:
        if field not in row:
            continue
        value = parse_number(row.get(field))
        if value is not None:
            car[field] = value
        elif clear_empty:
            car.pop(field, None)

    if parse_number(row.get("consumption_l_100km")) is not None:
        car.pop("consumption_kg_100km", None)
    if parse_number(row.get("consumption_kg_100km")) is not None:
        car.pop("consumption_l_100km", None)
    return car


def slim_car(car: dict[str, Any]) -> dict[str, Any]:
    slim: dict[str, Any] = {}
    slim["id"] = first_text(car, "id")
    slim["category"] = category_of(car)
    slim["brand"] = first_text(car, "brand")
    slim["model"] = first_text(car, "model")
    slim["version"] = first_text(car, "version", "powertrain")
    slim["fuel"] = fuel_of(car)

    source_url = first_text(car, "source_url", "motornet_detail_url")
    image_url = first_text(car, "image_url", "image_local_path", "image_source_url")
    if source_url:
        slim["source_url"] = source_url
    if image_url:
        slim["image_url"] = image_url

    for field in NUMERIC_FIELDS:
        value = first_number(car, field)
        if value is not None:
            slim[field] = value

    return {key: value for key, value in slim.items() if value not in (None, "")}


def build_json(rows: list[dict[str, Any]], base_payload: dict[str, Any], clear_empty: bool) -> dict[str, Any]:
    base_cars = [car for car in base_payload.get("cars", []) if isinstance(car, dict)]
    by_id = {clean(car.get("id")): car for car in base_cars if clean(car.get("id"))}

    out_cars: list[dict[str, Any]] = []
    seen: set[str] = set()

    for row in rows:
        car_id = clean(row.get("id"))
        if not car_id:
            continue
        base = by_id.get(car_id, {"id": car_id})
        merged = apply_excel_row(base, row, clear_empty=clear_empty)
        out_cars.append(slim_car(merged))
        seen.add(car_id)

    for car in base_cars:
        car_id = clean(car.get("id"))
        if car_id and car_id not in seen:
            out_cars.append(slim_car(car))

    return {
        "source": base_payload.get("source") or "motornet.it",
        "status": "ok",
        "schema": "cars_motornet_slim_v1",
        "cars": out_cars,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert cleaned Motornet Excel back to slim JSON.")
    parser.add_argument("--excel", required=True, help="Excel workbook exported by audit_motornet_quality.py")
    parser.add_argument("--sheet", default="catalog", help="Worksheet name")
    parser.add_argument("--base-json", default="data/cars_motornet.json", help="Original JSON used as base")
    parser.add_argument("--out", default="data/cars_motornet.cleaned.json", help="Output JSON path")
    parser.add_argument("--clear-empty", action="store_true", help="Allow empty Excel cells to clear JSON values")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    base_path = Path(args.base_json) if args.base_json else None
    out_path = Path(args.out)

    if not excel_path.exists():
        raise SystemExit(f"Excel not found: {excel_path}")
    if base_path is not None and not base_path.exists():
        raise SystemExit(f"Base JSON not found: {base_path}")

    rows = load_rows(excel_path, args.sheet)
    payload = build_json(rows, load_base_json(base_path), clear_empty=args.clear_empty)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Rows read: {len(rows)}")
    print(f"Cars written: {len(payload.get('cars', []))}")
    print(f"Output: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
