#!/usr/bin/env python3
import csv, html, io, json, re, statistics
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import unquote
import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PRICES = ROOT / 'data' / 'prices.json'
REGION_PAGE = 'https://www.mimit.gov.it/it/prezzo-medio-carburanti/regioni'
MIMIT_URLS = ['https://www.mimit.gov.it/images/exportCSV/prezzo_alle_8.csv','https://www.mise.gov.it/images/exportCSV/prezzo_alle_8.csv']
EU_PAGE = 'https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en'
REGIONS = ['Abruzzo','Basilicata','Calabria','Campania','Emilia Romagna','Friuli Venezia Giulia','Lazio','Liguria','Lombardia','Marche','Molise','Piemonte','Puglia','Sardegna','Sicilia','Toscana','Umbria',"Valle d'Aosta",'Veneto','Bolzano','Trento']
KEYS = {'benzina':['benzina'], 'gasolio':['gasolio','diesel'], 'gpl':['gpl'], 'metano':['metano']}
UNITS = {'benzina':'eur_l','gasolio':'eur_l','gpl':'eur_l','metano':'eur_kg'}
LABELS = {'benzina':'Benzina','gasolio':'Gasolio','gpl':'GPL','metano':'Metano'}

def load(p): return json.loads(p.read_text(encoding='utf-8')) if p.exists() else {}
def save(p,o): p.write_text(json.dumps(o, ensure_ascii=False, indent=2)+'\n', encoding='utf-8')
def cell(x): return str(x or '').replace('\ufeff','').strip()
def norm(x): return cell(x).lower().replace(' ','').replace('_','')
def val(x):
    try: return float(cell(x).replace(',','.'))
    except Exception: return None

def text_lines(markup):
    markup = re.sub(r'(?is)<script.*?</script>|<style.*?</style>', ' ', markup)
    markup = re.sub(r'(?i)</?(h\d|p|div|li|tr|td|th|br|section|article)[^>]*>', '\n', markup)
    markup = re.sub(r'<[^>]+>', ' ', markup)
    txt = html.unescape(markup)
    return [re.sub(r'\s+', ' ', x).strip() for x in txt.splitlines() if re.sub(r'\s+', ' ', x).strip()]

def mimit_regions():
    try:
        r = requests.get(REGION_PAGE, timeout=(6,20), headers={'User-Agent':'Mozilla/5.0 elettrica-tco'})
        r.raise_for_status()
        lines = text_lines(r.text)
        regions = {}
        current = None
        update = None
        for line in lines:
            mdate = re.search(r'Aggiornamento\s+(\d{2}-\d{2}-\d{4})', line)
            if mdate: update = mdate.group(1)
            if line in REGIONS:
                current = line
                regions.setdefault(current, {})
                continue
            if current:
                m = re.match(r'^(Gasolio|Benzina|GPL|Metano)\s+(SELF|SERVITO)\s+([0-9]+[\.,][0-9]+)', line, re.I)
                if m:
                    label = m.group(1).lower()
                    key = {'benzina':'benzina','gasolio':'gasolio','gpl':'gpl','metano':'metano'}[label]
                    regions[current][key] = val(m.group(3))
        regions = {k:v for k,v in regions.items() if any(v.get(f) for f in KEYS)}
        data = {}
        for k in KEYS:
            nums = [v[k] for v in regions.values() if v.get(k)]
            if nums: data[k] = round(statistics.fmean(nums), 3)
        if data.get('benzina') and data.get('gasolio'):
            data.update({'source':REGION_PAGE,'frequency':'daily_region_average','regions':regions,'samples':{k:len([v for v in regions.values() if v.get(k)]) for k in KEYS},'mimit_region_update':update,'average_method':'simple_mean_of_region_averages'})
            print('MIMIT regions parsed', {'regions':len(regions),'samples':data['samples'],'update':update})
            return data
        print('MIMIT regions failed: insufficient fuel values', {'regions':len(regions),'data':data})
    except Exception as e:
        print('MIMIT regions source failed', e)
    return {}

def find_col(cols, wants):
    for c in cols:
        n = norm(c)
        if any(norm(w) in n for w in wants): return c
    return None

def rows_after_header(text):
    try: dialect = csv.Sniffer().sniff(text[:5000], delimiters=';,|\t')
    except Exception:
        dialect = csv.excel; dialect.delimiter=';'
    rows = [[cell(c) for c in r] for r in csv.reader(io.StringIO(text), dialect=dialect)]
    rows = [r for r in rows if any(r)]
    for i,r in enumerate(rows):
        nr = [norm(c) for c in r]
        if any('carburante' in c or 'prodotto' in c for c in nr) and any('prezzo' in c for c in nr):
            out=[]
            for rr in rows[i+1:]:
                rr = (rr + ['']*len(r))[:len(r)]
                out.append(dict(zip(r, rr)))
            return r,out,i
    raise ValueError('header not found: '+repr(rows[:4]))

def fuel_key(label):
    t = str(label or '').lower()
    if any(x in t for x in ['special','premium','plus']): return None
    for k, aliases in KEYS.items():
        if any(a in t for a in aliases): return k
    return None

def mimit_csv():
    for url in MIMIT_URLS:
        try:
            r = requests.get(url, timeout=(6,20), headers={'User-Agent':'Mozilla/5.0 elettrica-tco'})
            r.raise_for_status()
            cols, rows, skipped = rows_after_header(r.content.decode('utf-8-sig', errors='replace'))
            fc, pc, sc = find_col(cols,['descCarburante','carburante','prodotto']), find_col(cols,['prezzo','price']), find_col(cols,['isSelf','self'])
            buckets = {k:[] for k in KEYS}
            for row in rows:
                k, price = fuel_key(row.get(fc,'')), val(row.get(pc,''))
                if not k or not price or price <= 0 or price > 10: continue
                if sc and str(row.get(sc,'')).lower() in ['0','false','servito'] and k in ['benzina','gasolio']: continue
                buckets[k].append(price)
            data = {k:round(statistics.fmean(v),3) for k,v in buckets.items() if v}
            print('MIMIT CSV parsed', {'url':url,'skipped_rows':skipped,'samples':{k:len(v) for k,v in buckets.items()}})
            if data.get('benzina') and data.get('gasolio'):
                data.update({'source':url,'frequency':'daily_csv_raw','samples':{k:len(v) for k,v in buckets.items()},'average_method':'raw_station_mean'})
                return data
        except Exception as e:
            print('MIMIT CSV source failed', url, e)
    return {}

def pick_eu_link(markup):
    links = re.findall(r'href=["\']([^"\']+)["\']', markup)
    candidates = []
    for l in [x.replace('&amp;','&') for x in links]:
        d = unquote(l).lower()
        if ('document/download' in d and ('xlsx' in d or 'spreadsheet' in d)) or '.xlsx' in d:
            candidates.append(l)
    def score(l):
        d = unquote(l).lower(); s=0
        if 'with+taxes' in d or 'with taxes' in d: s += 100
        if 'without+taxes' in d or 'without taxes' in d: s -= 100
        if 'duties' in d: s -= 25
        if 'latest' in d or 'weekly' in d: s += 10
        if 'price' in d: s += 5
        return s
    candidates = sorted(candidates, key=score, reverse=True)
    print('EU candidate links', len(candidates), candidates[:3])
    if not candidates: return None
    l = candidates[0]
    return 'https://energy.ec.europa.eu'+l if l.startswith('/') else l

def eu_weekly():
    try:
        page = requests.get(EU_PAGE, timeout=(6,20), headers={'User-Agent':'Mozilla/5.0 elettrica-tco'})
        page.raise_for_status()
        url = pick_eu_link(page.text)
        if not url:
            print('EU source failed: no xlsx/document download link found')
            return {}
        x = requests.get(url, timeout=(6,35), headers={'User-Agent':'Mozilla/5.0 elettrica-tco'}); x.raise_for_status()
        wb = load_workbook(io.BytesIO(x.content), data_only=True, read_only=True)
        candidates=[]; italy=0
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells=[cell(v) for v in row]
                if any(c.lower() in ['italy','italia','it'] for c in cells):
                    italy += 1
                    nums=[val(c) for c in cells]; nums=[n for n in nums if n and .5 <= n <= 3]
                    if len(nums) >= 2: candidates.append(nums)
        if not candidates:
            print('EU source failed: no Italy numeric row')
            return {}
        nums=candidates[0]
        data={'benzina':round(max(nums[:3]),3),'gasolio':round(sorted(nums[:4])[1] if len(nums)>=4 else nums[1],3),'source':url,'frequency':'weekly_eu','samples':{'italy_rows':italy},'average_method':'eu_italy_row'}
        print('EU weekly parsed', data)
        return data
    except Exception as e:
        print('EU source failed', e); return {}

def main():
    prices = load(PRICES) or {'fuel':{}, 'electricity':{'home':.30,'solar':.08}}
    prices.setdefault('fuel', {}).setdefault('units', UNITS)
    data = mimit_regions(); status='updated_mimit_regions_daily'
    if not data:
        data = mimit_csv(); status='updated_mimit_csv_daily'
    if not data:
        data = eu_weekly(); status='updated_eu_weekly_partial_keep_gpl_metano_previous'
    if data:
        for k in KEYS:
            if data.get(k): prices['fuel'][k]=data[k]
        prices['fuel'].update({'source':data.get('source'), 'frequency':data.get('frequency'), 'samples':data.get('samples',{}), 'units':UNITS, 'average_method':data.get('average_method')})
        if data.get('regions'): prices['fuel']['regions']=data['regions']
        if data.get('mimit_region_update'): prices['fuel']['mimit_region_update']=data['mimit_region_update']
        prices['status']=status
    else:
        prices['fuel']['units']=UNITS; prices['status']='fallback_previous_values'
    prices['updated_at']=datetime.now(timezone.utc).isoformat(); save(PRICES, prices); print(json.dumps(prices, ensure_ascii=False, indent=2))
if __name__ == '__main__': main()
